"""
scripts/03_extract_text.py

Reads contract_catalog.xlsx, extracts raw text and page maps from each contract.
- DOCX: python-docx
- PDF (text): pdfplumber (primary)
- PDF (image/scanned): pytesseract fallback if pdfplumber yields < OCR_MIN_TEXT_LENGTH chars

Outputs:
  outputs/contract_text/{ContractID}.txt
  outputs/page_maps/{ContractID}.json   — {"1": "page 1 text", "2": "page 2 text", ...}

Usage: python scripts/03_extract_text.py
"""
import sys
import json
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from logger import get_logger, write_summary_xlsx, _RUN_TIMESTAMP

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
import config

import openpyxl

logger = get_logger(__name__)


# ---------------------------------------------------------------------------
# Extraction helpers
# ---------------------------------------------------------------------------

def extract_docx(file_path: Path) -> tuple[str, dict]:
    """Extract text from a DOCX file. Returns (full_text, page_map)."""
    from docx import Document
    doc = Document(str(file_path))
    paragraphs = [p.text for p in doc.paragraphs]
    full_text = "\n".join(paragraphs)
    # DOCX has no native page boundaries — treat the whole document as page 1
    page_map = {"1": full_text}
    return full_text, page_map


def extract_pdf_text(file_path: Path) -> tuple[str, dict]:
    """Extract text from a PDF using pdfplumber. Returns (full_text, page_map)."""
    import pdfplumber
    page_map = {}
    pages_text = []
    with pdfplumber.open(str(file_path)) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            page_map[str(i)] = text
            pages_text.append(text)
    full_text = "\n".join(pages_text)
    return full_text, page_map


def extract_pdf_ocr(file_path: Path) -> tuple[str, dict]:
    """
    OCR fallback using pytesseract + pdf2image.
    Returns (full_text, page_map).
    """
    import pytesseract
    from pdf2image import convert_from_path

    pytesseract.pytesseract.tesseract_cmd = config.TESSERACT_PATH

    images = convert_from_path(str(file_path), dpi=200)
    page_map = {}
    pages_text = []
    for i, img in enumerate(images, start=1):
        text = pytesseract.image_to_string(img)
        page_map[str(i)] = text
        pages_text.append(text)
    full_text = "\n".join(pages_text)
    return full_text, page_map


def extract_text(file_path: Path) -> tuple[str, dict]:
    """
    Route extraction based on file type.
    PDF: try pdfplumber first; fall back to pytesseract if text yield is low.
    DOCX: use python-docx.
    Returns (full_text, page_map).
    """
    suffix = file_path.suffix.lower()

    if suffix == ".docx":
        return extract_docx(file_path)

    elif suffix == ".pdf":
        text, page_map = extract_pdf_text(file_path)
        if len(text.strip()) < config.OCR_MIN_TEXT_LENGTH:
            logger.warning(
                f"Low text yield on {file_path.name} "
                f"({len(text.strip())} chars < {config.OCR_MIN_TEXT_LENGTH}) -- falling back to OCR"
            )
            text, page_map = extract_pdf_ocr(file_path)
        return text, page_map

    else:
        raise ValueError(f"Unsupported file type: {suffix}")


# ---------------------------------------------------------------------------
# Catalog reader
# ---------------------------------------------------------------------------

def read_catalog() -> list[dict]:
    """Read contract_catalog.xlsx and return list of dicts."""
    if not config.CONTRACT_CATALOG_XLSX.exists():
        raise FileNotFoundError(
            f"contract_catalog.xlsx not found at {config.CONTRACT_CATALOG_XLSX}. "
            "Run 02_intake.py first."
        )
    wb = openpyxl.load_workbook(config.CONTRACT_CATALOG_XLSX)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return rows


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    summary_events = []
    logger.info("[STEP 1] Starting text extraction")

    catalog = read_catalog()
    if not catalog:
        logger.warning("contract_catalog.xlsx is empty — nothing to extract")
        write_summary_xlsx(summary_events)
        return

    logger.info(f"  {len(catalog)} contract(s) to process")

    config.CONTRACT_TEXT_DIR.mkdir(parents=True, exist_ok=True)
    config.PAGE_MAPS_DIR.mkdir(parents=True, exist_ok=True)

    success = 0
    failed = 0
    ocr_fallback = 0

    for row in catalog:
        cid = row["ContractID"]
        file_path = Path(row["FilePath"])
        txt_out = config.CONTRACT_TEXT_DIR / f"{cid}.txt"
        map_out = config.PAGE_MAPS_DIR / f"{cid}.json"

        if not file_path.exists():
            msg = f"{cid}: source file not found at {file_path}"
            logger.error(msg)
            summary_events.append({
                "run_timestamp": _RUN_TIMESTAMP,
                "script": "03_extract_text",
                "contract_id": cid,
                "level": "ERROR",
                "message": msg,
            })
            txt_out.write_text("EXTRACTION_FAILED", encoding="utf-8")
            map_out.write_text("{}", encoding="utf-8")
            failed += 1
            continue

        try:
            # Check if OCR fallback will be needed (for logging purposes)
            was_ocr = False
            suffix = file_path.suffix.lower()

            if suffix == ".pdf":
                # Peek at pdfplumber yield first to detect OCR path
                import pdfplumber
                with pdfplumber.open(str(file_path)) as pdf:
                    quick_text = "".join((p.extract_text() or "") for p in pdf.pages)
                if len(quick_text.strip()) < config.OCR_MIN_TEXT_LENGTH:
                    was_ocr = True
                    ocr_fallback += 1

            full_text, page_map = extract_text(file_path)

            txt_out.write_text(full_text, encoding="utf-8")
            map_out.write_text(json.dumps(page_map, ensure_ascii=False, indent=2), encoding="utf-8")

            char_count = len(full_text)
            pages = len(page_map)
            ocr_note = " [OCR]" if was_ocr else ""
            logger.info(f"  {cid}: {pages} page(s), {char_count} chars{ocr_note} <- {file_path.name}")
            success += 1

        except Exception as e:
            msg = f"{cid}: extraction failed — {e}"
            logger.error(msg)
            summary_events.append({
                "run_timestamp": _RUN_TIMESTAMP,
                "script": "03_extract_text",
                "contract_id": cid,
                "level": "ERROR",
                "message": msg,
            })
            txt_out.write_text("EXTRACTION_FAILED", encoding="utf-8")
            map_out.write_text("{}", encoding="utf-8")
            failed += 1

    logger.info(
        f"[STEP 1] Complete -- {success} extracted, {failed} failed, "
        f"{ocr_fallback} OCR fallback(s) -> {config.CONTRACT_TEXT_DIR}"
    )

    if failed > 0:
        summary_events.append({
            "run_timestamp": _RUN_TIMESTAMP,
            "script": "03_extract_text",
            "contract_id": "",
            "level": "WARNING",
            "message": f"{failed} contract(s) failed extraction — check individual ERROR entries above",
        })

    write_summary_xlsx(summary_events)


if __name__ == "__main__":
    main()
