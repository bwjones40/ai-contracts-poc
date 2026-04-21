"""
scripts/06_summarize.py

LiteLLM contract summarization — generates plain-language summaries for each contract.
Reads contract_text/*.txt and extracted_fields.xlsx.
Writes outputs/summaries.xlsx.

Usage: python scripts/06_summarize.py
"""

import sys
import json
import os
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
import litellm

from logger import get_logger, write_summary_xlsx, _RUN_TIMESTAMP

# Config
PROJECT_ROOT = Path(__file__).parent.parent
OUTPUTS_DIR = PROJECT_ROOT / "outputs"
CONTRACT_TEXT_DIR = OUTPUTS_DIR / "contract_text"
CONTRACT_CATALOG_XLSX = OUTPUTS_DIR / "contract_catalog.xlsx"
EXTRACTED_FIELDS_XLSX = OUTPUTS_DIR / "extracted_fields.xlsx"
SUMMARIES_XLSX = OUTPUTS_DIR / "summaries.xlsx"

LITELLM_MODEL = os.environ.get("LITELLM_MODEL", "gpt-4o")
LITELLM_API_BASE = os.environ.get("LITELLM_API_BASE", "http://localhost:4000")
LITELLM_API_KEY = os.environ.get("LITELLM_API_KEY", "")
MAX_TOKENS_SUMMARY = 500
SUMMARY_WORD_LIMIT = 6000  # truncate input text if very long

logger = get_logger(__name__)

SYSTEM_PROMPT = """You are a contract review assistant helping category managers quickly understand contracts.
Write a 3-5 sentence plain-language summary. Focus on: parties, what the contract covers,
key dates, value, and notable terms. Write for a business reader, not a lawyer.
Do not fabricate details. If something is unclear, say so.
Return only the summary text — no headers, bullets, or JSON."""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_catalog() -> list[dict]:
    wb = openpyxl.load_workbook(CONTRACT_CATALOG_XLSX)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def load_key_fields() -> dict[str, dict]:
    """Returns {contract_id: {field_name: value}} for key fields only."""
    wb = openpyxl.load_workbook(EXTRACTED_FIELDS_XLSX)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    contracts: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        cid = r["ContractID"]
        if cid not in contracts:
            contracts[cid] = {}
        contracts[cid][r["FieldName"]] = r.get("ExtractedValue")
    return contracts


def read_contract_text(contract_id: str) -> str:
    txt_path = CONTRACT_TEXT_DIR / f"{contract_id}.txt"
    if not txt_path.exists():
        return ""
    text = txt_path.read_text(encoding="utf-8", errors="replace")
    # Truncate to word limit
    words = text.split()
    if len(words) > SUMMARY_WORD_LIMIT:
        text = " ".join(words[:SUMMARY_WORD_LIMIT]) + "\n[... truncated ...]"
    return text


def call_llm_summary(contract_id: str, contract_text: str, key_fields: dict) -> str:
    """Call LiteLLM to generate a plain-language summary. Returns summary string. Retries once on None response."""
    import time

    field_context = "\n".join(
        f"  {k}: {v}" for k, v in key_fields.items() if v and str(v).strip()
    )
    user_content = f"""Contract ID: {contract_id}

Key extracted fields:
{field_context}

Contract text:
{contract_text}

Write a 3-5 sentence plain-language summary of this contract."""

    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": user_content},
    ]

    for attempt in range(2):
        response = litellm.completion(
            model=LITELLM_MODEL,
            messages=messages,
            api_base=LITELLM_API_BASE,
            api_key=LITELLM_API_KEY,
            max_tokens=MAX_TOKENS_SUMMARY,
        )
        content = response.choices[0].message.content
        if content is not None:
            return content.strip()
        if attempt == 0:
            logger.warning(f"  {contract_id}: LLM returned None content on attempt 1 — retrying")
            time.sleep(2)

    raise ValueError(f"LLM returned None content on both attempts for {contract_id}")


def detect_contract_type(key_fields: dict) -> str:
    """Return ContractType from extracted fields, or 'Unknown'."""
    ct = key_fields.get("ContractType")
    if ct and str(ct).strip() and str(ct).strip().upper() not in ("NULL", "NONE", ""):
        return str(ct).strip()
    return "Unknown"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run_summarize():
    logger.info("[STEP 0] Starting summarization — 06_summarize.py")
    summary_events = []

    catalog = load_catalog()
    all_fields = load_key_fields()
    logger.info(f"[STEP 1] Loaded catalog: {len(catalog)} contracts")

    summaries = []
    succeeded = 0
    failed = 0

    for record in catalog:
        cid = record["ContractID"]
        filename = record["FileName"]
        logger.info(f"[STEP 2] Summarizing {cid} ({filename})")

        contract_text = read_contract_text(cid)
        if not contract_text:
            logger.warning(f"  No text found for {cid} — skipping")
            summary_events.append({
                "script": "06_summarize",
                "contract_id": cid,
                "level": "WARNING",
                "message": f"No contract text file found for {cid}",
            })
            failed += 1
            continue

        key_fields = all_fields.get(cid, {})
        contract_type = detect_contract_type(key_fields)

        try:
            summary_text = call_llm_summary(cid, contract_text, key_fields)
            summaries.append({
                "ContractID": cid,
                "FileName": filename,
                "Summary": summary_text,
                "ContractType": contract_type,
                "SummaryTimestamp": datetime.now().isoformat(),
            })
            succeeded += 1
            logger.info(f"  {cid} summarized ({len(summary_text)} chars)")

        except Exception as e:
            err_msg = f"Summary failed for {cid}: {e}"
            logger.error(f"  {err_msg}")
            summary_events.append({
                "script": "06_summarize",
                "contract_id": cid,
                "level": "ERROR",
                "message": err_msg,
            })
            summaries.append({
                "ContractID": cid,
                "FileName": filename,
                "Summary": "SUMMARY_FAILED",
                "ContractType": contract_type,
                "SummaryTimestamp": datetime.now().isoformat(),
            })
            failed += 1

    logger.info(f"[STEP 3] Summarization complete -- {succeeded} succeeded, {failed} failed")

    # Write summaries.xlsx
    logger.info("[STEP 4] Writing summaries.xlsx")
    OUTPUTS_DIR.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summaries"

    col_headers = ["ContractID", "FileName", "Summary", "ContractType", "SummaryTimestamp"]
    ws.append(col_headers)

    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    for s in summaries:
        ws.append([s[h] for h in col_headers])

    # Wrap text in Summary column (column C = index 3)
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # Set row height for summary rows
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 80

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 22

    ws.freeze_panes = "A2"

    wb.save(SUMMARIES_XLSX)
    logger.info(f"[STEP 4] Complete -- {len(summaries)} rows -> {SUMMARIES_XLSX}")

    write_summary_xlsx(summary_events)
    logger.info("[STEP 5] 06_summarize.py complete")

    return summaries


if __name__ == "__main__":
    run_summarize()
