"""
scripts/07_build_validation.py

Assembles validation_review.xlsx from extracted_fields, risk_signals, and summaries.
Primary human review artifact for the AI Contracts POC.

Usage: python scripts/07_build_validation.py
"""

import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

from logger import get_logger, write_summary_xlsx

PROJECT_ROOT = Path(__file__).parent.parent
OUTPUTS_DIR = PROJECT_ROOT / "outputs"
EXTRACTED_FIELDS_XLSX = OUTPUTS_DIR / "extracted_fields.xlsx"
RISK_SIGNALS_XLSX = OUTPUTS_DIR / "risk_signals.xlsx"
SUMMARIES_XLSX = OUTPUTS_DIR / "summaries.xlsx"
VALIDATION_XLSX = OUTPUTS_DIR / "validation_review.xlsx"
CONFIDENCE_THRESHOLD_WARN = 0.6

logger = get_logger(__name__)

SEVERITY_ORDER = {"High": 3, "Medium": 2, "Low": 1, "": 0}

# Column positions in Review sheet (1-indexed)
COL_CONTRACT_ID     = 1   # A
COL_FILE_NAME       = 2   # B
COL_FIELD_NAME      = 3   # C
COL_EXTRACTED       = 4   # D
COL_CONFIDENCE      = 5   # E
COL_EVIDENCE_PAGE   = 6   # F
COL_EVIDENCE_SNIP   = 7   # G
COL_SEVERITY        = 8   # H
COL_RISK_MSG        = 9   # I
COL_OVERRIDE        = 10  # J
COL_CHANGE_REASON   = 11  # K
COL_APPROVED        = 12  # L
COL_REVIEWER        = 13  # M
COL_REVIEW_TS       = 14  # N
COL_FINAL_VALUE     = 15  # O

HEADERS = [
    "ContractID", "FileName", "FieldName", "ExtractedValue", "Confidence",
    "EvidencePage", "EvidenceSnippet", "SeverityFlag", "RiskMessage",
    "ReviewerOverride", "ChangeReason", "Approved", "Reviewer",
    "ReviewTimestamp", "FinalValue"
]

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_extracted_fields() -> list[dict]:
    wb = openpyxl.load_workbook(EXTRACTED_FIELDS_XLSX)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    logger.info(f"[STEP 1] Loaded {len(rows)} extracted field rows")
    return rows


def load_risk_signals() -> dict:
    """Returns {(contract_id, field_triggered): highest_severity_signal}"""
    if not RISK_SIGNALS_XLSX.exists():
        logger.warning("[STEP 2] risk_signals.xlsx not found — no risk data will be joined")
        return {}
    wb = openpyxl.load_workbook(RISK_SIGNALS_XLSX)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    signals: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        key = (r["ContractID"], r["FieldTriggered"])
        existing = signals.get(key)
        if existing is None or SEVERITY_ORDER.get(r["SeverityTier"], 0) > SEVERITY_ORDER.get(existing["SeverityTier"], 0):
            signals[key] = r
    logger.info(f"[STEP 2] Loaded risk signals for {len(signals)} (contract, field) pairs")
    return signals


def load_summaries() -> list[dict]:
    if not SUMMARIES_XLSX.exists():
        logger.warning("[STEP 3] summaries.xlsx not found — Summaries sheet will be empty")
        return []
    wb = openpyxl.load_workbook(SUMMARIES_XLSX)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    logger.info(f"[STEP 3] Loaded {len(rows)} summaries")
    return rows


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_review_sheet(ws, review_rows: list[dict]) -> int:
    """Write Review sheet. Returns data row count."""
    # Header
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Data rows
    for i, row in enumerate(review_rows, start=2):
        ws.cell(i, COL_CONTRACT_ID).value   = row["ContractID"]
        ws.cell(i, COL_FILE_NAME).value     = row["FileName"]
        ws.cell(i, COL_FIELD_NAME).value    = row["FieldName"]
        ws.cell(i, COL_EXTRACTED).value     = row["ExtractedValue"]
        ws.cell(i, COL_CONFIDENCE).value    = row["Confidence"]
        ws.cell(i, COL_EVIDENCE_PAGE).value = row["EvidencePage"]
        ws.cell(i, COL_EVIDENCE_SNIP).value = row["EvidenceSnippet"]
        ws.cell(i, COL_SEVERITY).value      = row["SeverityFlag"]
        ws.cell(i, COL_RISK_MSG).value      = row["RiskMessage"]
        # Columns J-N left blank (reviewer fills)
        # FinalValue formula: if override non-empty use it, else use extracted
        ws.cell(i, COL_FINAL_VALUE).value = f'=IF(J{i}<>"",J{i},D{i})'

    data_row_count = len(review_rows)
    last_row = data_row_count + 1  # +1 for header

    # Conditional formatting (formula-based, anchored on first data row)
    data_range = f"A2:O{last_row}"

    # Red: SeverityFlag = "High"  (evaluated first — wins over yellow on same row)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    red_font = Font(color="FFFFFF")
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=['$H2="High"'], fill=red_fill, font=red_font)
    )

    # Yellow: SeverityFlag = "Medium" OR Confidence < threshold
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'OR($H2="Medium",AND($E2<>"", $E2<{CONFIDENCE_THRESHOLD_WARN}))'],
            fill=yellow_fill
        )
    )

    # Green: Approved = "YES"
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(formula=['$L2="YES"'], fill=green_fill)
    )

    # Dropdown on Approved column (L)
    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
    dv.sqref = f"L2:L{last_row}"
    ws.add_data_validation(dv)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    return data_row_count


def build_summaries_sheet(ws, summaries: list[dict]):
    if not summaries:
        ws.append(["No summaries available"])
        return
    sum_headers = list(summaries[0].keys())
    ws.append(sum_headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in summaries:
        ws.append([row.get(h) for h in sum_headers])
    # Auto-fit — cap Summary column wider
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)
    ws.freeze_panes = "A2"


def build_risk_signals_sheet(ws):
    if not RISK_SIGNALS_XLSX.exists():
        ws.append(["risk_signals.xlsx not found"])
        return
    wb_risk = openpyxl.load_workbook(RISK_SIGNALS_XLSX)
    ws_src = wb_risk.active
    for row_idx, row in enumerate(ws_src.iter_rows(values_only=True), start=1):
        ws.append(list(row))
        if row_idx == 1:
            for cell in ws[1]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_validation():
    logger.info("[STEP 0] Starting 07_build_validation.py")
    summary_events = []

    fields = load_extracted_fields()
    signals = load_risk_signals()
    summaries = load_summaries()

    # Build review rows: join extracted fields with highest-severity risk signal per (contract, field)
    review_rows = []
    for f in fields:
        cid = f.get("ContractID", "")
        fname = f.get("FieldName", "")
        signal = signals.get((cid, fname))

        review_rows.append({
            "ContractID":    cid,
            "FileName":      f.get("FileName", ""),
            "FieldName":     fname,
            "ExtractedValue": f.get("ExtractedValue", ""),
            "Confidence":    f.get("Confidence"),
            "EvidencePage":  f.get("EvidencePage"),
            "EvidenceSnippet": f.get("EvidenceSnippet", ""),
            "SeverityFlag":  signal["SeverityTier"] if signal else "",
            "RiskMessage":   signal["Message"] if signal else "",
        })

    logger.info(f"[STEP 4] Built {len(review_rows)} review rows")

    # Create workbook
    OUTPUTS_DIR.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()

    # Review sheet (active)
    ws_review = wb.active
    ws_review.title = "Review"
    n_rows = build_review_sheet(ws_review, review_rows)
    logger.info(f"[STEP 5] Review sheet complete — {n_rows} data rows")

    # Summaries sheet
    ws_sum = wb.create_sheet("Summaries")
    build_summaries_sheet(ws_sum, summaries)
    logger.info(f"[STEP 6] Summaries sheet added — {len(summaries)} rows")

    # Risk Signals sheet (read-only copy)
    ws_risk = wb.create_sheet("Risk Signals")
    build_risk_signals_sheet(ws_risk)
    logger.info("[STEP 7] Risk Signals sheet added")

    wb.save(VALIDATION_XLSX)
    logger.info(f"[STEP 8] Complete -- {n_rows} rows -> {VALIDATION_XLSX}")

    write_summary_xlsx(summary_events)
    logger.info("[STEP 9] 07_build_validation.py complete")
    return n_rows


if __name__ == "__main__":
    build_validation()
