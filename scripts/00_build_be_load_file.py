"""
scripts/00_build_be_load_file.py

Generates the Coupa Business Entity Load File from source_business_entities.xlsx.
This is a PREREQUISITE step — run before the main pipeline.

Output: outputs/Target_Business_Entity_Load_File.xlsx

Usage: python scripts/00_build_be_load_file.py
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font

from logger import get_logger, write_summary_xlsx

PROJECT_ROOT = Path(__file__).parent.parent
SOURCE_DATA_DIR = PROJECT_ROOT / "source_data"
OUTPUTS_DIR = PROJECT_ROOT / "outputs"
SOURCE_BE_XLSX = SOURCE_DATA_DIR / "source_business_entities.xlsx"
BE_LOAD_FILE_XLSX = OUTPUTS_DIR / "Target_Business_Entity_Load_File.xlsx"

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Header rows — exact spec, do not deviate
# ---------------------------------------------------------------------------

HEADER_ROW_1 = [
    "Business Entity", "ID", "Name", "Display Name", "Type", "Status",
    "Country of Origin Code", "State of Origin", "Formation Type",
    "Business Entity Alternate Names", "Parent Business Entity Id"
]
HEADER_ROW_2 = [
    "Contact", "ID", "Primary", "Email", "Name Given", "Name Family",
    "Phone Work", "Phone Mobile", "Phone Fax"
]
HEADER_ROW_3 = [
    "Address", "Id", "Primary", "Name", "Line 1", "Line 2", "Line 3",
    "Line 4", "City", "State", "Postal Code", "Country Code", "Plant"
]
HEADER_ROW_4 = [
    "Business Entity External Reference", "ID", "Name", "Type", "Value"
]
HEADER_ROW_5 = [
    "Supplier Sharing Setting", "ID", "Sharable ID", "Sharable Type",
    "Shared Category", "Share with Children"
]

HEADER_ROWS = [HEADER_ROW_1, HEADER_ROW_2, HEADER_ROW_3, HEADER_ROW_4, HEADER_ROW_5]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_str(value) -> str:
    """Strip whitespace from value and return as string. Returns '' for None."""
    if value is None:
        return ""
    return str(value).strip()


def write_text_row(ws, row_data: list):
    """Append a row, forcing text format on all cells to prevent Excel auto-conversion."""
    ws.append([safe_str(v) for v in row_data])
    row_idx = ws.max_row  # row we just appended
    for col_idx in range(1, len(row_data) + 1):
        ws.cell(row_idx, col_idx).number_format = "@"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_be_load_file():
    logger.info("[STEP 0] Starting 00_build_be_load_file.py")
    summary_events = []

    # Validate source file exists
    if not SOURCE_BE_XLSX.exists():
        msg = (
            f"Source file not found: {SOURCE_BE_XLSX}\n"
            "  Place source_business_entities.xlsx in the source_data/ folder and re-run."
        )
        logger.error(f"[STEP 1] {msg}")
        summary_events.append({
            "script": "00_build_be_load_file",
            "contract_id": "",
            "level": "ERROR",
            "message": msg,
        })
        write_summary_xlsx(summary_events)
        sys.exit(1)

    # Load source
    wb_src = openpyxl.load_workbook(SOURCE_BE_XLSX, data_only=True)
    ws_src = wb_src.active
    src_headers = [ws_src.cell(1, c).value for c in range(1, ws_src.max_column + 1)]
    logger.info(f"[STEP 1] Source loaded — {ws_src.max_row - 1} data rows, columns: {src_headers}")

    # Build case-insensitive column index (partial match)
    def find_col(keyword: str) -> int | None:
        kw = keyword.lower()
        for i, h in enumerate(src_headers):
            if h and kw in str(h).lower():
                return i
        return None

    col_idx = {
        "Name":         find_col("name"),
        "AlternateName":find_col("alternate name"),
        "Type":         find_col("type"),
        "Status":       find_col("status"),
        "FormationType":find_col("formation type"),
        "Line1":        find_col("line 1"),
        "Line2":        find_col("line 2"),
        "Line3":        find_col("line 3"),
        "Line4":        find_col("line 4"),
        "City":         find_col("city"),
        "State":        find_col("state"),
        "PostalCode":   find_col("postal code"),
        "CountryCode":  find_col("country code"),
        "ExtRefName1":  find_col("external reference name 1"),
        "ExtRefValue1": find_col("external reference value 1"),
        "ExtRefName2":  find_col("external reference name 2"),
        "ExtRefValue2": find_col("external reference value 2"),
    }
    logger.info(f"[STEP 1] Column mapping resolved: {col_idx}")

    def get(row_values, key: str) -> str:
        idx = col_idx.get(key)
        if idx is None or idx >= len(row_values):
            return ""
        return safe_str(row_values[idx])

    # Build target workbook
    OUTPUTS_DIR.mkdir(exist_ok=True)
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Target"

    # Write 5 header rows (plain text, bold)
    header_font = Font(bold=True)
    for header_row in HEADER_ROWS:
        ws_out.append(header_row)
    for row_idx in range(1, 6):
        for cell in ws_out[row_idx]:
            cell.font = header_font
    logger.info("[STEP 2] 5 header rows written")

    # Process source data rows
    source_rows = list(ws_src.iter_rows(min_row=2, values_only=True))
    n_source = len(source_rows)
    counts = {"be": 0, "contact": 0, "address": 0, "ext_ref": 0, "warnings": 0}
    expected_data_rows = 0

    for row_num_0, row_values in enumerate(source_rows, start=0):
        row_num = row_num_0 + 2  # 1-indexed source row number (row 1 = header)

        name   = get(row_values, "Name")
        type_  = get(row_values, "Type")
        status = get(row_values, "Status")

        # Warn on missing required fields
        for field_label, field_val in [("Name", name), ("Type", type_), ("Status", status)]:
            if not field_val:
                msg = f"Source row {row_num}: {field_label} is blank — row included but may be rejected by Coupa"
                logger.warning(f"  {msg}")
                summary_events.append({
                    "script": "00_build_be_load_file",
                    "contract_id": f"source_row_{row_num}",
                    "level": "WARNING",
                    "message": msg,
                })
                counts["warnings"] += 1

        # Business Entity row
        write_text_row(ws_out, [
            "Business Entity",            # A
            "",                           # B — ID
            name,                         # C — Name
            "",                           # D — Display Name
            type_,                        # E — Type
            status,                       # F — Status
            "",                           # G — Country of Origin Code
            "",                           # H — State of Origin
            get(row_values, "FormationType"),  # I — Formation Type
            get(row_values, "AlternateName"),  # J — Alternate Names
            "",                           # K — Parent Business Entity Id
        ])
        counts["be"] += 1
        expected_data_rows += 1

        # Contact row (placeholder — all fields blank except literal)
        write_text_row(ws_out, ["Contact"] + [""] * 8)
        counts["contact"] += 1
        expected_data_rows += 1

        # Address row
        write_text_row(ws_out, [
            "Address",                    # A
            "",                           # B — Id
            "",                           # C — Primary
            "",                           # D — Name
            get(row_values, "Line1"),     # E — Line 1
            get(row_values, "Line2"),     # F — Line 2
            get(row_values, "Line3"),     # G — Line 3
            get(row_values, "Line4"),     # H — Line 4
            get(row_values, "City"),      # I — City
            get(row_values, "State"),     # J — State
            get(row_values, "PostalCode"),# K — Postal Code
            get(row_values, "CountryCode"),# L — Country Code
            "",                           # M — Plant
        ])
        counts["address"] += 1
        expected_data_rows += 1

        # Ext Ref row 1 (conditional)
        ext_ref_name_1  = get(row_values, "ExtRefName1")
        ext_ref_value_1 = get(row_values, "ExtRefValue1")
        if ext_ref_name_1 or ext_ref_value_1:
            write_text_row(ws_out, [
                "Business Entity External Reference",
                "", ext_ref_name_1, "", ext_ref_value_1
            ])
            counts["ext_ref"] += 1
            expected_data_rows += 1

        # Ext Ref row 2 (conditional)
        ext_ref_name_2  = get(row_values, "ExtRefName2")
        ext_ref_value_2 = get(row_values, "ExtRefValue2")
        if ext_ref_name_2 or ext_ref_value_2:
            write_text_row(ws_out, [
                "Business Entity External Reference",
                "", ext_ref_name_2, "", ext_ref_value_2
            ])
            counts["ext_ref"] += 1
            expected_data_rows += 1

    logger.info(
        f"[STEP 3] Processed {n_source} source rows -> "
        f"BE: {counts['be']}, Contact: {counts['contact']}, "
        f"Address: {counts['address']}, ExtRef: {counts['ext_ref']}, "
        f"Warnings: {counts['warnings']}"
    )

    # Validate total row count
    expected_total = 5 + expected_data_rows  # 5 header rows + data rows
    actual_rows = ws_out.max_row
    assert actual_rows == expected_total, (
        f"Row count mismatch: expected {expected_total}, got {actual_rows}"
    )
    logger.info(f"[STEP 4] Row count validation passed — {actual_rows} total rows ({expected_data_rows} data)")

    # Auto-fit columns
    col_widths: dict = {}
    for row in ws_out.iter_rows():
        for cell in row:
            val_len = len(str(cell.value or ""))
            ltr = cell.column_letter
            if val_len > col_widths.get(ltr, 0):
                col_widths[ltr] = val_len
    for ltr, width in col_widths.items():
        ws_out.column_dimensions[ltr].width = min(width + 2, 50)

    wb_out.save(BE_LOAD_FILE_XLSX)
    logger.info(f"[STEP 5] Complete -- {actual_rows} rows -> {BE_LOAD_FILE_XLSX}")

    write_summary_xlsx(summary_events)
    logger.info("[STEP 6] 00_build_be_load_file.py complete")
    return actual_rows


if __name__ == "__main__":
    build_be_load_file()
