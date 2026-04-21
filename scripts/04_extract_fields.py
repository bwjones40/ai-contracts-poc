"""
scripts/04_extract_fields.py

Reads contract_catalog.xlsx + page_maps/*.json, calls LiteLLM for structured
field extraction per schema.json, and writes outputs/extracted_fields.xlsx
(long format: one row per field per contract).

Usage: python scripts/04_extract_fields.py
"""
import sys
import json
import re
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

from logger import get_logger, write_summary_xlsx, _RUN_TIMESTAMP

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
import config

import litellm
import openpyxl

logger = get_logger(__name__)

SCRIPT_NAME = "04_extract_fields"
MAX_WORDS = 8_000

SYSTEM_PROMPT = """
You are a contract data extraction assistant.
Rules:
1. Only extract values explicitly stated in the contract text. Do not infer or fabricate.
2. If a field cannot be found, return null for value and explain in null_reason.
3. For each extracted value, identify the page number and a short quote (under 100 characters).
4. Confidence 0.0-1.0. Be conservative -- only use 0.9+ for values clearly and explicitly stated.
5. Return ONLY valid JSON. No markdown, no preamble, no explanation outside the JSON.

Output format -- JSON array, one object per field:
[
  {
    "field_name": "Supplier",
    "extracted_value": "Apex Industrial Services",
    "confidence": 0.95,
    "evidence_page": 1,
    "evidence_snippet": "entered into by Apex Industrial Services (Supplier)",
    "is_null": false,
    "null_reason": null
  }
]
"""

REPAIR_PROMPT = """
Your previous response was not valid JSON. Return ONLY the JSON array — no markdown fences,
no preamble, no explanation. Start with [ and end with ].
"""

OUTPUT_HEADERS = [
    "ContractID",
    "FileName",
    "FieldName",
    "ExtractedValue",
    "Confidence",
    "EvidencePage",
    "EvidenceSnippet",
    "IsNull",
    "NullReason",
    "ExtractionTimestamp",
]


def load_schema(schema_path: Path) -> list[dict]:
    with open(schema_path, encoding="utf-8") as f:
        data = json.load(f)
    return data["fields"]


def load_catalog(catalog_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(catalog_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        if record.get("ContractID"):
            rows.append(record)
    return rows


def build_page_marked_text(page_map: dict) -> str:
    parts = []
    for page_num in sorted(page_map.keys(), key=lambda x: int(x)):
        parts.append(f"--- PAGE {page_num} ---")
        parts.append(page_map[page_num])
    return "\n".join(parts)


def truncate_to_words(text: str, max_words: int) -> tuple[str, bool]:
    words = text.split()
    if len(words) <= max_words:
        return text, False
    return " ".join(words[:max_words]), True


def strip_json_fences(raw: str) -> str:
    """Remove markdown code fences if the model wrapped the JSON in them."""
    raw = raw.strip()
    # Remove ```json ... ``` or ``` ... ```
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return raw.strip()


def call_llm(messages: list[dict]) -> str:
    response = litellm.completion(
        model=config.LITELLM_MODEL,
        messages=messages,
        api_base=config.LITELLM_API_BASE,
        api_key=config.LITELLM_API_KEY,
        max_tokens=config.MAX_TOKENS_EXTRACTION,
    )
    return response.choices[0].message.content


def extract_fields_for_contract(
    contract_id: str,
    file_name: str,
    page_map: dict,
    field_names: list[str],
    summary_events: list[dict],
) -> list[dict]:
    """
    Run extraction for one contract. Returns a list of row dicts.
    On unrecoverable failure, returns EXTRACTION_FAILED rows for all fields.
    """
    ts = datetime.now().isoformat(timespec="seconds")

    page_text, truncated = truncate_to_words(build_page_marked_text(page_map), MAX_WORDS)
    if truncated:
        msg = f"Text truncated to {MAX_WORDS} words"
        logger.warning(f"[{contract_id}] {msg}")
        summary_events.append({
            "run_timestamp": _RUN_TIMESTAMP,
            "script": SCRIPT_NAME,
            "contract_id": contract_id,
            "level": "WARNING",
            "message": msg,
        })

    user_content = (
        f"Extract the following fields from this contract:\n"
        f"{json.dumps(field_names)}\n\n"
        f"{page_text}"
    )

    messages = [
        {"role": "system", "content": SYSTEM_PROMPT.strip()},
        {"role": "user", "content": user_content},
    ]

    raw = None
    parsed = None

    # Attempt 1
    try:
        raw = call_llm(messages)
        parsed = json.loads(strip_json_fences(raw))
    except Exception as e:
        logger.warning(f"[{contract_id}] Attempt 1 failed ({type(e).__name__}: {e}) — retrying")

    # Attempt 2 (repair prompt)
    if parsed is None:
        try:
            repair_messages = messages + [
                {"role": "assistant", "content": raw or ""},
                {"role": "user", "content": REPAIR_PROMPT.strip()},
            ]
            raw2 = call_llm(repair_messages)
            parsed = json.loads(strip_json_fences(raw2))
        except Exception as e:
            msg = f"Extraction failed after retry ({type(e).__name__}: {e})"
            logger.error(f"[{contract_id}] {msg}")
            summary_events.append({
                "run_timestamp": _RUN_TIMESTAMP,
                "script": SCRIPT_NAME,
                "contract_id": contract_id,
                "level": "ERROR",
                "message": msg,
            })
            return _failed_rows(contract_id, file_name, field_names, ts)

    # Build result rows from parsed JSON
    parsed_by_field = {item.get("field_name"): item for item in parsed if isinstance(item, dict)}
    rows = []
    for field_name in field_names:
        item = parsed_by_field.get(field_name, {})
        confidence = item.get("confidence")
        extracted_value = item.get("extracted_value") or ""
        is_null = item.get("is_null", True)
        null_reason = item.get("null_reason") or ""
        evidence_page = item.get("evidence_page")
        evidence_snippet = (item.get("evidence_snippet") or "")[:100]

        if confidence is not None and float(confidence) < config.CONFIDENCE_THRESHOLD_WARN:
            msg = f"Low confidence ({confidence:.2f}) on field '{field_name}'"
            logger.warning(f"[{contract_id}] {msg}")
            summary_events.append({
                "run_timestamp": _RUN_TIMESTAMP,
                "script": SCRIPT_NAME,
                "contract_id": contract_id,
                "level": "WARNING",
                "message": msg,
            })

        rows.append({
            "ContractID": contract_id,
            "FileName": file_name,
            "FieldName": field_name,
            "ExtractedValue": extracted_value,
            "Confidence": confidence,
            "EvidencePage": evidence_page,
            "EvidenceSnippet": evidence_snippet,
            "IsNull": is_null,
            "NullReason": null_reason,
            "ExtractionTimestamp": ts,
        })

    return rows


def _failed_rows(contract_id: str, file_name: str, field_names: list[str], ts: str) -> list[dict]:
    return [
        {
            "ContractID": contract_id,
            "FileName": file_name,
            "FieldName": fn,
            "ExtractedValue": "EXTRACTION_FAILED",
            "Confidence": None,
            "EvidencePage": None,
            "EvidenceSnippet": "",
            "IsNull": True,
            "NullReason": "Extraction failed — see pipeline_log_summary.xlsx",
            "ExtractionTimestamp": ts,
        }
        for fn in field_names
    ]


def write_extracted_fields(all_rows: list[dict], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExtractedFields"
    ws.append(OUTPUT_HEADERS)
    for row in all_rows:
        ws.append([row[h] for h in OUTPUT_HEADERS])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    summary_events: list[dict] = []

    logger.info("[STEP 1] Starting — loading schema and catalog")
    schema_fields = load_schema(ROOT / "schema.json")
    field_names = [f["name"] for f in schema_fields]
    logger.info(f"[STEP 1] Complete — {len(field_names)} fields in schema")

    catalog = load_catalog(config.CONTRACT_CATALOG_XLSX)
    logger.info(f"[STEP 2] Starting — {len(catalog)} contracts to process")

    all_rows: list[dict] = []
    ok_count = 0
    fail_count = 0

    for i, record in enumerate(catalog, 1):
        contract_id = record["ContractID"]
        file_name = record.get("FileName", "")
        page_map_path = config.PAGE_MAPS_DIR / f"{contract_id}.json"

        if not page_map_path.exists():
            msg = f"page_map not found — skipping"
            logger.error(f"[{contract_id}] {msg}")
            summary_events.append({
                "run_timestamp": _RUN_TIMESTAMP,
                "script": SCRIPT_NAME,
                "contract_id": contract_id,
                "level": "ERROR",
                "message": msg,
            })
            all_rows.extend(_failed_rows(contract_id, file_name, field_names,
                                         datetime.now().isoformat(timespec="seconds")))
            fail_count += 1
            continue

        with open(page_map_path, encoding="utf-8") as f:
            page_map = json.load(f)

        logger.info(f"[{i}/{len(catalog)}] Extracting fields for {contract_id}")
        rows = extract_fields_for_contract(
            contract_id, file_name, page_map, field_names, summary_events
        )
        all_rows.extend(rows)

        if any(r["ExtractedValue"] == "EXTRACTION_FAILED" for r in rows):
            fail_count += 1
        else:
            ok_count += 1

    logger.info(f"[STEP 2] Complete — {ok_count} succeeded, {fail_count} failed")

    logger.info("[STEP 3] Writing extracted_fields.xlsx")
    write_extracted_fields(all_rows, config.EXTRACTED_FIELDS_XLSX)
    total_rows = len(all_rows)
    logger.info(
        f"[STEP 3] Complete — {total_rows} rows written to {config.EXTRACTED_FIELDS_XLSX}"
    )

    write_summary_xlsx(summary_events)
    if summary_events:
        logger.info(f"Summary log updated — {len(summary_events)} WARNING/ERROR events")


if __name__ == "__main__":
    main()
