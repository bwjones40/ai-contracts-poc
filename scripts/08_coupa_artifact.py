"""
scripts/08_coupa_artifact.py

Generates coupa_ready.xlsx from rows in validation_review.xlsx where Approved == "YES".
Only approved contracts are included — this gate is enforced in code, not by convention.

Usage: python scripts/08_coupa_artifact.py
"""

import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from logger import get_logger, write_summary_xlsx

PROJECT_ROOT = Path(__file__).parent.parent
OUTPUTS_DIR = PROJECT_ROOT / "outputs"
VALIDATION_XLSX = OUTPUTS_DIR / "validation_review.xlsx"
COUPA_READY_XLSX = OUTPUTS_DIR / "coupa_ready.xlsx"

logger = get_logger(__name__)

COUPA_COLUMNS = [
    "contract_name", "supplier_name", "contract_type", "effective_date",
    "expiration_date", "contract_value", "business_entity", "governing_law",
    "auto_renewal", "payment_terms", "source_file", "extracted_by",
    "approved_by", "approval_date", "coupa_upload_status"
]

# FieldName in extracted_fields -> coupa_ready column
FIELD_TO_COUPA = {
    "Supplier":         "supplier_name",
    "ContractType":     "contract_type",
    "EffectiveDate":    "effective_date",
    "ExpirationDate":   "expiration_date",
    "ContractValue":    "contract_value",
    "BusinessEntity":   "business_entity",
    "GoverningLaw":     "governing_law",
    "AutoRenewal":      "auto_renewal",
    "PaymentTerms":     "payment_terms",
}


# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------

def load_approved_contracts() -> tuple[dict, dict]:
    """
    Reads validation_review.xlsx Review sheet.
    Returns:
        approved: {contract_id: {field_name: final_value}}
        meta:     {contract_id: {source_file, approved_by, approval_date}}

    FinalValue is computed as: ReviewerOverride if non-empty, else ExtractedValue.
    (Avoids dependency on Excel formula cache.)
    """
    if not VALIDATION_XLSX.exists():
        raise FileNotFoundError(
            f"validation_review.xlsx not found at {VALIDATION_XLSX}. "
            "Run 07_build_validation.py first."
        )

    wb = openpyxl.load_workbook(VALIDATION_XLSX, data_only=True)
    ws = wb["Review"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers)}

    approved: dict = {}
    meta: dict = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        approved_val = str(row[idx["Approved"]] or "").strip().upper()
        if approved_val != "YES":
            continue

        cid        = row[idx["ContractID"]]
        field_name = row[idx["FieldName"]]

        extracted_value   = row[idx["ExtractedValue"]] or ""
        reviewer_override = row[idx["ReviewerOverride"]] or ""
        final_value = reviewer_override if str(reviewer_override).strip() else extracted_value

        if cid not in approved:
            approved[cid] = {}
        approved[cid][field_name] = str(final_value) if final_value is not None else ""

        if cid not in meta:
            meta[cid] = {
                "source_file":   row[idx["FileName"]] or "",
                "approved_by":   row[idx["Reviewer"]] or "",
                "approval_date": str(row[idx["ReviewTimestamp"]] or ""),
            }

    logger.info(f"[STEP 1] Loaded {len(approved)} approved contracts from validation_review.xlsx")
    return approved, meta


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

def build_coupa_row(cid: str, fields: dict, meta: dict) -> dict:
    row = {col: "" for col in COUPA_COLUMNS}

    supplier = fields.get("Supplier", "")
    ctype    = fields.get("ContractType", "")
    row["contract_name"]       = f"{supplier} {ctype} {cid}".strip()
    row["extracted_by"]        = "AI-POC-v1"
    row["coupa_upload_status"] = "PENDING"
    row["source_file"]         = meta.get("source_file", "")
    row["approved_by"]         = meta.get("approved_by", "")
    row["approval_date"]       = meta.get("approval_date", "")

    for field_name, coupa_col in FIELD_TO_COUPA.items():
        row[coupa_col] = fields.get(field_name, "")

    return row


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_coupa():
    logger.info("[STEP 0] Starting 08_coupa_artifact.py")
    summary_events = []

    approved, meta = load_approved_contracts()

    if not approved:
        logger.warning(
            "[STEP 2] No approved contracts found — "
            "coupa_ready.xlsx will contain header only. "
            "Open validation_review.xlsx, set Approved=YES on at least one row, then re-run."
        )
        summary_events.append({
            "script": "08_coupa_artifact",
            "contract_id": "",
            "level": "WARNING",
            "message": "No rows with Approved=YES found in validation_review.xlsx"
        })

    coupa_rows = [build_coupa_row(cid, approved[cid], meta.get(cid, {})) for cid in sorted(approved)]
    logger.info(f"[STEP 3] Built {len(coupa_rows)} Coupa rows")

    # Build workbook
    OUTPUTS_DIR.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CoupaPayload"

    # Warning banner — row 1
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    banner_text = (
        f"Generated: {ts} | By: AI-POC-v1 | Status: PENDING — "
        "HUMAN REVIEW REQUIRED BEFORE UPLOAD"
    )
    ws.append([banner_text])
    banner_cell = ws.cell(1, 1)
    banner_cell.font = Font(bold=True, color="FFFFFF")
    banner_cell.fill = PatternFill("solid", fgColor="C00000")
    banner_cell.alignment = Alignment(horizontal="left")

    # Merge banner across all columns
    last_col_letter = chr(ord("A") + len(COUPA_COLUMNS) - 1)
    ws.merge_cells(f"A1:{last_col_letter}1")

    # Header row — row 2
    ws.append(COUPA_COLUMNS)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    # Data rows — start at row 3
    for row in coupa_rows:
        ws.append([str(row.get(c, "") or "") for c in COUPA_COLUMNS])

    # Auto-fit columns — use get_column_letter to avoid merged cell issues
    for col_idx in range(1, len(COUPA_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(r, col_idx).value or "")) for r in range(2, ws.max_row + 1)),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A3"

    wb.save(COUPA_READY_XLSX)
    logger.info(f"[STEP 4] Complete -- {len(coupa_rows)} rows -> {COUPA_READY_XLSX}")

    write_summary_xlsx(summary_events)
    logger.info("[STEP 5] 08_coupa_artifact.py complete")
    return len(coupa_rows)


if __name__ == "__main__":
    build_coupa()
