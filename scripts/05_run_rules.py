"""
scripts/05_run_rules.py

Rules engine — evaluates extracted fields against rules.xlsx and writes risk_signals.xlsx.
Only evaluates rules where Enabled == "YES".

Usage: python scripts/05_run_rules.py
"""

import sys
import os
from pathlib import Path
from datetime import date, datetime

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from dateutil import parser as dateutil_parser

from logger import get_logger, write_summary_xlsx, _RUN_TIMESTAMP

# Config — resolve from project root regardless of working directory
PROJECT_ROOT = Path(__file__).parent.parent
OUTPUTS_DIR = PROJECT_ROOT / "outputs"
EXTRACTED_FIELDS_XLSX = OUTPUTS_DIR / "extracted_fields.xlsx"
RISK_SIGNALS_XLSX = OUTPUTS_DIR / "risk_signals.xlsx"
RULES_XLSX = PROJECT_ROOT / "rules.xlsx"
CONFIDENCE_THRESHOLD_WARN = 0.6

logger = get_logger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_rules() -> list[dict]:
    """Load enabled rules from rules.xlsx."""
    wb = openpyxl.load_workbook(RULES_XLSX)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rules = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        if str(r.get("Enabled", "")).strip().upper() == "YES":
            rules.append(r)
    logger.info(f"[STEP 1] Loaded {len(rules)} enabled rules from rules.xlsx")
    return rules


def load_extracted_fields() -> dict[str, dict[str, dict]]:
    """
    Returns nested dict: {contract_id: {field_name: {value, confidence, ...}}}
    """
    wb = openpyxl.load_workbook(EXTRACTED_FIELDS_XLSX)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    contracts: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        cid = r["ContractID"]
        fname = r["FieldName"]
        if cid not in contracts:
            contracts[cid] = {}
        contracts[cid][fname] = {
            "value": r.get("ExtractedValue"),
            "confidence": r.get("Confidence"),
            "is_null": r.get("IsNull"),
            "null_reason": r.get("NullReason"),
        }
    logger.info(f"[STEP 2] Loaded fields for {len(contracts)} contracts from extracted_fields.xlsx")
    return contracts


def safe_parse_date(value) -> date | None:
    """Parse a date string to date object. Returns None if unparseable."""
    if not value:
        return None
    s = str(value).strip()
    if not s or s.upper() in ("TBD", "N/A", "UNKNOWN", "NULL", "NONE", "EXTRACTION_FAILED"):
        return None
    try:
        return dateutil_parser.parse(s, fuzzy=True).date()
    except Exception:
        return None


def is_null_value(value) -> bool:
    """True if the value is effectively missing/null."""
    if value is None:
        return True
    s = str(value).strip()
    return s == "" or s.upper() in ("NULL", "NONE", "TBD", "N/A", "UNKNOWN", "EXTRACTION_FAILED")


def build_signal(rule: dict, contract_id: str, evidence: str) -> dict:
    return {
        "ContractID": contract_id,
        "RuleID": rule["RuleID"],
        "SeverityTier": rule["Severity"],
        "Message": rule["MessageTemplate"],
        "Evidence": str(evidence) if evidence is not None else "",
        "FieldTriggered": rule["Target"],
        "RuleTimestamp": datetime.now().isoformat(),
    }


# ---------------------------------------------------------------------------
# Rule evaluation
# ---------------------------------------------------------------------------

def evaluate_rule(rule: dict, field_data: dict | None, confidence: float | None, contract_id: str) -> dict | None:
    """
    Evaluate a single rule against a single contract's field data.
    Returns a signal dict if the rule fires, None otherwise.
    """
    rt = rule["RuleType"]
    cond = rule["Condition"]
    target = rule["Target"]

    field_value = field_data.get("value") if field_data else None

    if rt == "missing_field" and cond == "is_null":
        if is_null_value(field_value):
            return build_signal(rule, contract_id, field_value)

    elif rt == "date_check":
        parsed = safe_parse_date(field_value)
        if parsed:
            if cond == "past_due" and parsed < date.today():
                return build_signal(rule, contract_id, field_value)
            if cond == "within_90_days":
                days = (parsed - date.today()).days
                if 0 <= days <= 90:
                    return build_signal(rule, contract_id, field_value)
        # If date is unparseable and rule is date_check, skip silently (not a signal)

    elif rt == "clause_check":
        normalized = str(field_value or "").strip().upper()
        if cond == "present":
            # Fire if value indicates presence: "PRESENT" exactly, or any non-empty value
            # that isn't a negative indicator (LLM sometimes returns descriptive text instead of "PRESENT")
            if normalized and normalized not in ("NOT_FOUND", "NULL", "NONE", ""):
                return build_signal(rule, contract_id, field_value)
        if cond == "missing" and normalized in ("", "NOT_FOUND", "NULL", "NONE"):
            return build_signal(rule, contract_id, field_value)

    elif rt == "value_check" and cond == "below_threshold" and target == "Confidence":
        # R010: fire if ANY field for this contract has confidence below threshold
        # This is evaluated per-field in the caller; confidence passed directly
        if confidence is not None and float(confidence) < CONFIDENCE_THRESHOLD_WARN:
            return build_signal(rule, contract_id, f"confidence={confidence:.2f}")

    return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run_rules():
    logger.info("[STEP 0] Starting rules engine — 05_run_rules.py")
    summary_events = []

    rules = load_rules()
    contracts = load_extracted_fields()

    signals = []
    r010_rule = next((r for r in rules if r["RuleID"] == "R010"), None)
    standard_rules = [r for r in rules if r["RuleID"] != "R010"]

    for cid, fields in contracts.items():
        # Evaluate standard rules (field-level)
        for rule in standard_rules:
            target = rule["Target"]
            field_data = fields.get(target)
            confidence = field_data.get("confidence") if field_data else None
            signal = evaluate_rule(rule, field_data, confidence, cid)
            if signal:
                signals.append(signal)
                logger.debug(
                    f"  Rule {rule['RuleID']} fired on {cid} / {target} "
                    f"(severity={rule['Severity']})"
                )

        # R010: fire once per field that has low confidence (not null fields)
        if r010_rule:
            for fname, fdata in fields.items():
                conf = fdata.get("confidence")
                val = fdata.get("value")
                is_null = fdata.get("is_null")
                # Only flag if field returned a value but confidence is low
                if not is_null and not is_null_value(val) and conf is not None and float(conf) < CONFIDENCE_THRESHOLD_WARN:
                    sig = {
                        "ContractID": cid,
                        "RuleID": r010_rule["RuleID"],
                        "SeverityTier": r010_rule["Severity"],
                        "Message": r010_rule["MessageTemplate"],
                        "Evidence": f"{fname}: confidence={float(conf):.2f}",
                        "FieldTriggered": fname,
                        "RuleTimestamp": datetime.now().isoformat(),
                    }
                    signals.append(sig)
                    logger.debug(f"  R010 fired on {cid} / {fname} (confidence={conf:.2f})")

    logger.info(f"[STEP 3] Rule evaluation complete — {len(signals)} signals generated across {len(contracts)} contracts")

    # Write risk_signals.xlsx
    logger.info("[STEP 4] Writing risk_signals.xlsx")
    OUTPUTS_DIR.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Signals"

    headers = ["ContractID", "RuleID", "SeverityTier", "Message", "Evidence", "FieldTriggered", "RuleTimestamp"]
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    severity_colors = {"High": "FF0000", "Medium": "FFC000", "Low": "92D050"}

    for sig in signals:
        row_idx = ws.max_row + 1
        ws.append([sig[h] for h in headers])
        sev = sig.get("SeverityTier", "")
        color = severity_colors.get(sev)
        if color:
            for col in range(1, len(headers) + 1):
                ws.cell(row_idx, col).fill = PatternFill("solid", fgColor=color)

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"

    wb.save(RISK_SIGNALS_XLSX)
    logger.info(f"[STEP 4] Complete -- {len(signals)} risk signals -> {RISK_SIGNALS_XLSX}")

    # Log summary of which rules fired
    rule_counts: dict[str, int] = {}
    for sig in signals:
        rid = sig["RuleID"]
        rule_counts[rid] = rule_counts.get(rid, 0) + 1
    for rid, cnt in sorted(rule_counts.items()):
        logger.info(f"  {rid}: {cnt} signals")

    # Collect summary events for any contracts where no rules fired (informational only)
    write_summary_xlsx(summary_events)
    logger.info("[STEP 5] 05_run_rules.py complete")

    return signals


if __name__ == "__main__":
    run_rules()
